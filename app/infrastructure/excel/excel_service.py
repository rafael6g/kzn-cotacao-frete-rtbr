import pandas as pd
from pathlib import Path
from typing import Optional
from datetime import datetime, timedelta, timezone

from app.domain.entities.cotacao import Cotacao, StatusCotacao
from app.core.logging_config import get_logger

logger = get_logger(__name__)

# Colunas fixas adicionadas ao Excel de saída (fretes são colunas dinâmicas)
COLUNAS_RESULTADO = [
    "resultado_origem",
    "resultado_destino",
    "resultado_status",
    "resultado_fonte",
    "resultado_tempo_viagem",
    "resultado_distancia_km",
    "resultado_rota_descricao",
    "resultado_valor_pedagio",
    "resultado_valor_combustivel",
    "resultado_valor_total",
    "resultado_consultado_em",
    "resultado_erro",
]


class ExcelService:

    def __init__(self, settings):
        self._settings = settings

    def ler_arquivo(self, caminho: str) -> list[dict]:
        """
        Lê Excel ou CSV e retorna lista de dicts (uma por linha).
        Colunas esperadas mínimas: origem, destino (case-insensitive).
        """
        path = Path(caminho)
        suffix = path.suffix.lower()

        try:
            if suffix in (".xlsx", ".xls"):
                df = pd.read_excel(path, dtype=str)
            elif suffix == ".csv":
                df = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
            else:
                raise ValueError(f"Formato não suportado: {suffix}")
        except Exception as e:
            from app.domain.exceptions import ExcelInvalidoError
            raise ExcelInvalidoError(f"Erro ao ler arquivo: {e}")

        # Normaliza nomes de colunas (lower + strip)
        df.columns = [c.lower().strip() for c in df.columns]
        df = df.fillna("")

        # Valida colunas obrigatórias
        obrigatorias = {"origem", "destino"}
        faltando = obrigatorias - set(df.columns)
        if faltando:
            from app.domain.exceptions import ExcelInvalidoError
            raise ExcelInvalidoError(
                f"Colunas obrigatórias ausentes: {faltando}. "
                f"Colunas encontradas: {list(df.columns)}"
            )

        return df.to_dict("records")

    async def gerar(
        self,
        arquivo_entrada: str,
        cotacoes: list[Cotacao],
        output_path: str,
        validade_cache_horas: int = 0,
    ) -> None:
        """
        Gera Excel de resultado:
        - Mantém todas as colunas originais
        - Adiciona colunas de resultado
        - Ordena por linha_numero
        """
        path = Path(arquivo_entrada) if arquivo_entrada else None
        suffix = path.suffix.lower() if path else ""

        if path and path.is_file():
            try:
                if suffix in (".xlsx", ".xls"):
                    df_original = pd.read_excel(path, dtype=str)
                else:
                    df_original = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
            except Exception as e:
                logger.error(f"Erro ao reler arquivo original: {e}")
                df_original = pd.DataFrame()
        else:
            df_original = pd.DataFrame()

        # Constrói DataFrame de resultados
        cotacoes_sorted = sorted(cotacoes, key=lambda c: c.linha_numero)
        rows = []
        for cotacao in cotacoes_sorted:
            row = {"_linha": cotacao.linha_numero}
            r = cotacao.resultado

            row["resultado_origem_site"] = cotacao.parametros.origem
            row["resultado_destino_site"] = cotacao.parametros.destino
            row["resultado_param_site"] = cotacao.parametros.site or "rotasbrasil"
            row["resultado_param_veiculo"] = cotacao.parametros.veiculo_label
            row["resultado_param_eixos"] = cotacao.parametros.eixos
            row["resultado_param_preco_combustivel"] = cotacao.parametros.preco_combustivel
            row["resultado_param_consumo_km_l"] = cotacao.parametros.consumo_km_l
            row["resultado_param_tabela_frete"] = getattr(cotacao.parametros, "tabela_frete", "A")
            row["resultado_param_retorno_vazio"] = getattr(cotacao.parametros, "retorno_vazio", False)
            row["resultado_status"] = cotacao.status.value
            row["resultado_fonte"] = cotacao.fonte.value if cotacao.fonte else ""
            row["resultado_tempo_viagem"] = r.tempo_viagem if r else ""
            row["resultado_distancia_km"] = r.distancia_km if r else ""
            row["resultado_rota_descricao"] = r.rota_descricao if r else ""
            row["resultado_valor_pedagio"] = r.valor_pedagio if r else ""
            row["resultado_valor_combustivel"] = r.valor_combustivel if r else ""
            row["resultado_valor_total"] = r.valor_total if r else ""
            consultado_em_fmt = ""
            if r and r.consultado_em:
                try:
                    dt_c = datetime.fromisoformat(r.consultado_em)
                    consultado_em_fmt = dt_c.strftime("%d/%m/%Y %H:%M")
                except Exception:
                    consultado_em_fmt = r.consultado_em
            row["resultado_consultado_em"] = consultado_em_fmt
            # Validade do cache
            validade_ate = ""
            if r and r.consultado_em and validade_cache_horas > 0:
                try:
                    dt = datetime.fromisoformat(r.consultado_em)
                    validade_ate = (dt + timedelta(hours=validade_cache_horas)).strftime("%d/%m/%Y %H:%M")
                except Exception:
                    pass
            row["resultado_validade_ate"] = validade_ate
            # Colunas dinâmicas — um campo por tipo de carga (ex: "Tipo_Carga_Carga Geral")
            for tipo, val in (r.fretes.items() if r else {}.items()):
                row[tipo] = val
            # Praças de pedágio — formato: "P3 - Jacarezinho | R$64,00 (12,80 eixo) | BR-369 - KM 1.500"
            if r and r.pedagios:
                row["resultado_pedagios"] = "\n".join(
                    f"{p['nome']} | {p['tarifa']} ({p['por_eixo']}) | {p['rodovia']}"
                    for p in r.pedagios
                )
            else:
                row["resultado_pedagios"] = ""
            row["resultado_erro"] = cotacao.erro_mensagem or ""
            rows.append(row)

        df_resultado = pd.DataFrame(rows)

        # Mescla original + resultado pelo índice de linha
        if not df_original.empty and not df_resultado.empty:
            df_resultado = df_resultado.rename(columns={"_linha": "_idx"})
            df_original.index = range(len(df_original))
            df_resultado.index = df_resultado["_idx"].apply(lambda x: x - 1)
            df_resultado = df_resultado.drop(columns=["_idx"])
            df_final = df_original.join(df_resultado, how="left")
        else:
            df_resultado = df_resultado.drop(columns=["_linha"], errors="ignore")
            df_final = df_resultado

        # Remove prefixo "resultado_" das colunas de saída
        rename = {c: c[len("resultado_"):] for c in df_final.columns if c.startswith("resultado_")}
        df_final = df_final.rename(columns=rename)

        # Renomeia colunas originais de entrada e do scraper para nomes finais legíveis
        rename_final = {}
        if "origem" in df_final.columns:
            rename_final["origem"] = "origem excel"
        if "destino" in df_final.columns:
            rename_final["destino"] = "destino excel"
        if "origem_site" in df_final.columns:
            rename_final["origem_site"] = "origem site"
        if "destino_site" in df_final.columns:
            rename_final["destino_site"] = "destino site"
        if rename_final:
            df_final = df_final.rename(columns=rename_final)

        # Define a ordem exata das colunas de resultado
        ORDEM_RESULTADO = [
            "origem excel",
            "destino excel",
            "origem site",
            "destino site",
            "param_site",
            "param_veiculo",
            "param_eixos",
            "param_preco_combustivel",
            "param_consumo_km_l",
            "param_tabela_frete",
            "param_retorno_vazio",
            "tempo_viagem",
            "distancia_km",
            "rota_descricao",
            "valor_pedagio",
            "valor_combustivel",
            "valor_total",
            "antt_ccd",
            "antt_cc",
            "antt_valor_ida",
            "antt_valor_retorno",
            "Tipo_Carga_Granel Sólido",
            "Tipo_Carga_Granel Líquido",
            "Tipo_Carga_Frigorificada",
            "Tipo_Carga_Conteinerizada",
            "Tipo_Carga_Carga Geral",
            "Tipo_Carga_Neogranel",
            "Tipo_Carga_Perigosa (granel sólido)",
            "Tipo_Carga_Perigosa (granel líquido)",
            "Tipo_Carga_Perigosa (frigorificada)",
            "Tipo_Carga_Perigosa (conteinerizada)",
            "Tipo_Carga_Perigosa (carga geral)",
            "Tipo_Carga_Granel Pressurizada",
            "pedagios",
            "status",
            "fonte",
            "consultado_em",
            "validade_ate",
            "erro",
        ]

        # Colunas extras do arquivo de entrada (exceto as já incluídas na ORDEM_RESULTADO)
        colunas_entrada_extras = [
            c for c in df_final.columns
            if c not in ORDEM_RESULTADO
        ]
        # Ordem final: resultado ordenado + colunas extras do cliente no fim
        colunas_presentes = [
            c for c in ORDEM_RESULTADO if c in df_final.columns
        ] + colunas_entrada_extras
        df_final = df_final.reindex(columns=colunas_presentes)

        # Estiliza e salva
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_final.to_excel(writer, index=False, sheet_name="Cotações")
            self._estilizar(writer, df_final)

        logger.info(f"Excel salvo em: {output_path}")

    @staticmethod
    def _estilizar(writer, df: pd.DataFrame) -> None:
        """Aplica formatação: cabeçalho azul (entrada), vermelho Real94 (resultado)."""
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            ws = writer.sheets["Cotações"]

            # Azul = colunas de entrada originais (origem/destino excel)
            # Vermelho Real94 = colunas de resultado do scraper
            azul_escuro = PatternFill("solid", fgColor="1565C0")
            vermelho    = PatternFill("solid", fgColor="BF181B")
            branco      = Font(color="FFFFFF", bold=True)

            COLUNAS_RESULTADO = {
                "origem site", "destino site",
                "param_site", "param_veiculo", "param_eixos",
                "param_preco_combustivel", "param_consumo_km_l", "param_tabela_frete",
                "param_retorno_vazio",
                "status", "fonte", "tempo_viagem",
                "distancia_km", "rota_descricao", "valor_pedagio",
                "valor_combustivel", "valor_total",
                "antt_ccd", "antt_cc", "antt_valor_ida", "antt_valor_retorno",
                "consultado_em", "validade_ate", "erro", "pedagios",
            }
            for cell in ws[1]:
                col = str(cell.value or "")
                is_resultado = col in COLUNAS_RESULTADO or col.startswith("Tipo_Carga_")
                cell.fill = vermelho if is_resultado else azul_escuro
                cell.font = branco
                cell.alignment = Alignment(horizontal="center")

            # Auto-largura de colunas
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        except Exception as e:
            logger.warning(f"Não foi possível estilizar o Excel: {e}")
